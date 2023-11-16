import argparse
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, NamedStyle
import csv
# import pandas as pd
# from openpyxl.utils.dataframe import dataframe_to_rows
from extractXMLsProcessor import XMLFilesProcessor

class ExcelFileProcessor:
    def __init__(self, csv_folder, output_file):
        # Initialize the workbook creator with the directory of CSV files and the output file path
        self.csv_folder = csv_folder
        self.output_file = output_file
        self.wb = Workbook()
        self.toc = self.wb.active
        self._setup_toc()

    def _setup_toc(self):
        # Set up the Table of Contents sheet
        self.toc.title = "TOC"
        # Create TOC heading style
        heading_style = NamedStyle(name="heading", font=Font(bold=True), alignment=Alignment(horizontal="center"))
        self.wb.add_named_style(heading_style)
        self.toc['A1'] = "Table of Contents"
        self.toc['A1'].style = heading_style
        self.toc.column_dimensions[get_column_letter(1)].width = 20

    def _format_worksheet(self, ws):
        # Apply bold font to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Adjust the column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = max_length

    def _add_csv_to_sheet(self, csv_file):
        # Add each CSV file to a new sheet in the workbook
        # Truncate sheet title to a maximum of 31 characters for Excel compatibility
        sheet_title = os.path.splitext(csv_file)[0]
        ws = self.wb.create_sheet(title=sheet_title[:31])
        csv_path = os.path.join(self.csv_folder, csv_file)
        
        # Solution 1 - Using CSV module from python
        # Read the CSV file and add its contents to the sheet
        try:
            with open(csv_path, mode='r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter='~')
                for row in reader:
                    ws.append(row)
        except UnicodeDecodeError:
            # If utf-8 decoding fails, try another encoding such as 'latin1'
            with open(csv_path, mode='r', encoding='latin1') as f:
                reader = csv.reader(f, delimiter='~')
                for row in reader:
                    ws.append(row)

        # Solution 2 - Using pandas - can be used if Solution 1 has perf issue .. 
        # try:
        #     # df = pd.read_csv(csv_path, sep='~', encoding='utf-8-sig')
        #     # Set keep_default_na to False to avoid treating 'None' as NaN, Specify na_values to an empty list to ensure only empty strings are treated as NaN
        #     df = pd.read_csv(csv_path, sep='~', encoding='utf-8', keep_default_na=False, na_values=[])
        #     for row in dataframe_to_rows(df, index=False, header=True):
        #         ws.append(row)
        # except UnicodeDecodeError:
        #     try:
        #         df = pd.read_csv(csv_path, sep='~', encoding='latin1', keep_default_na=False, na_values=[])
        #         for row in dataframe_to_rows(df, index=False, header=True):
        #             ws.append(row)
        #     except Exception as e:
        #         raise e
        # except pd.errors.ParserError as e:
        #     raise e   
        
        # Call the new method to apply formatting
        self._format_worksheet(ws)

    def _add_to_toc(self, sheet_title, index):
        # Add the sheet name to the TOC with a hyperlink
        toc_cell = self.toc.cell(column=1, row=index+1, value=sheet_title)
        # Truncate sheet title to a maximum of 31 characters for Excel compatibility
        toc_cell.hyperlink = f"#{sheet_title[:31]}!A1"
        toc_cell.style = 'Hyperlink'

    def _remove_existing_output_file(self):
        # Check if the output file already exists and remove it if it does
        if os.path.exists(self.output_file):
            os.remove(self.output_file)
            print(f"Removed existing file: {self.output_file}")

    def create_excel_with_toc(self):
        # Create the workbook and save it to the output file
        self._remove_existing_output_file()
        try:
            if list(filter(lambda f: f.endswith('.csv'), os.listdir(self.csv_folder))):
                for index, csv_file in enumerate(sorted(os.listdir(self.csv_folder)), start=2):
                    if csv_file.endswith('.csv'):
                        self._add_csv_to_sheet(csv_file)
                        self._add_to_toc(os.path.splitext(csv_file)[0], index)
                self.wb.save(self.output_file)
                print(f"Excel file saved at {self.output_file}")
            else:
                print('No *.csv files found for : '+self.output_file+' - File not created !')
        except Exception as e:
            print(f"Failed to create Excel file: {e}")

    def run():
        # Set up command line argument parsing
        parser = argparse.ArgumentParser(description="Generate CSV files from XML Files. Generate Excel workbook from CSV files with TOC.")
        parser.add_argument('-i', '--input_dir', required=True, help='Input directory containing XML files.')
        parser.add_argument('-o', '--output_dir', required=True, help='Output directory for CSV and Excel files.')

        args = parser.parse_args()

        # Create the output directory if it doesn't exist
        try:
            os.makedirs(args.output_dir, exist_ok=True)
        except OSError as e:
            print(f"Failed to create output directory: {e}")
            return

        # Iterate over all files in output directory
        print("All existing *.csv files in output directory will be removed.")
        for filename in os.listdir(args.output_dir):
            if filename.endswith('.csv'):
                os.remove(os.path.join(args.output_dir, filename))

        # Convert XML files to CSV
        xml_processor = XMLFilesProcessor(args.input_dir, args.output_dir)
        xml_processor.process_files()

        # Define the name for the output Excel file and create the workbook
        excel_output_file = os.path.join(args.output_dir, os.path.basename(os.path.normpath(args.output_dir))+'.xlsx')
        creator = ExcelFileProcessor(args.output_dir, excel_output_file)
        creator.create_excel_with_toc()

if __name__ == "__main__":
    ExcelFileProcessor.run()
