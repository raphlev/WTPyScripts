import os
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "PowerPoint Document Data"

# Write the header row
header = ['Document', 'Directory', 'N# Pages', 'Size']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\SUPPORTS_FORMATIONS_INDIGO_OPALE'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

# Initialize the PowerPoint application
def initialize_powerpoint():
    try:
        ppt_app = win32com.client.Dispatch('PowerPoint.Application')
        # Remove or comment out the following line
        # ppt_app.Visible = False  # PowerPoint does not support hiding the application window
        ppt_app.Visible = True  # PowerPoint requires the application to be visible - if False is used, the application raises an error
        return ppt_app
    except Exception as e:
        print(f"Error initializing PowerPoint application: {e}")
        exit(1)

ppt_app = initialize_powerpoint()

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.ppt', '.pptx')):
            file_path = os.path.join(root, file)

            # Increment the file counter before processing
            file_count += 1

            print(f"\nProcessing file {file_count}: {file_path}")
            print(f"File name: {file}")

            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            directory_path = os.path.dirname(file_path)
            size = ''
            num_slides = ''

            try:
                # Get the size of the document
                try:
                    size = os.path.getsize(file_path)
                    print(f"Size of '{file_path}': {size} bytes")
                except Exception as e:
                    print(f"Error getting size for '{file_path}': {e}")
                    size = ''

                # Open the presentation
                try:
                    presentation = ppt_app.Presentations.Open(file_path, WithWindow=False)
                    print(f"Successfully opened presentation '{file_path}'")
                except Exception as e:
                    print(f"Error opening '{file_path}': {e}")
                    skipped_files.append(file_path)
                    continue

                # Get the total number of slides
                try:
                    num_slides = presentation.Slides.Count
                    print(f"Total slides in '{file_path}': {num_slides}")
                except Exception as e:
                    print(f"Error getting slide count for '{file_path}': {e}")
                    num_slides = ''

                # Write the data to Excel
                ws.append([
                    file_name,
                    directory_path,
                    num_slides,
                    size
                ])
                print(f"Data written to Excel for '{file_name}'.")

            except Exception as e:
                print(f"Error processing '{file_path}': {e}")
                skipped_files.append(file_path)
            finally:
                # Ensure the presentation is closed
                try:
                    if presentation is not None:
                        print(f"Attempting to close presentation '{file_path}'")
                        presentation.Close()
                        del presentation
                        print(f"Successfully closed presentation '{file_path}'")
                except Exception as e:
                    print(f"Error closing presentation '{file_path}': {e}")
                    ppt_app = initialize_powerpoint()

print(f"\nTotal files processed: {file_count}")

# Close the PowerPoint application
try:
    ppt_app.Quit()
    ppt_app = None  # Release the COM object
    print("PowerPoint application closed.")
except Exception as e:
    print(f"Error quitting PowerPoint application: {e}")

# Save the Excel workbook
try:
    wb.save('ppt_output.xlsx')
    print("Excel workbook 'ppt_output.xlsx' saved successfully.")
except Exception as e:
    print(f"Error saving Excel workbook: {e}")

# Display the skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files were processed successfully.")
