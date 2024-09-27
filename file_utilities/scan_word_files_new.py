import os
import re
import time
import gc
import functools
import pywintypes
import win32com.client
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def remove_illegal_characters(value):
    """
    Removes illegal control characters from a string, except for allowed ones.
    
    Args:
        value (str): The string to sanitize.
    
    Returns:
        str: The sanitized string.
    """
    if isinstance(value, str):
        # Remove illegal characters (control characters except for allowed ones)
        value = re.sub(
            r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]",
            "",
            value
        )
    return value

# Retry decorator for COM calls with enhanced logging
def retry_on_COM_error(max_retries=3, delay=1):
    """
    Decorator to retry a function upon encountering a COM error.
    
    Args:
        max_retries (int): Maximum number of retries.
        delay (int): Delay in seconds between retries.
    
    Returns:
        function: The decorated function with retry logic.
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            while retries <= max_retries:
                try:
                    return func(*args, **kwargs)
                except pywintypes.com_error as e:
                    if e.args[0] == -2147418111:  # Call was rejected by callee
                        retries += 1
                        print(f"COM call to '{func.__name__}' failed with 'Call was rejected by callee', retrying ({retries}/{max_retries})...")
                        time.sleep(delay)
                    else:
                        print(f"COM error in '{func.__name__}': {e}")
                        raise
                except Exception as e:
                    print(f"Unexpected error in '{func.__name__}': {e}")
                    raise
            print(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
            raise Exception(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
        return wrapper
    return decorator

# Initialize the Word application
def initialize_word():
    """
    Initializes and returns the Word application object.
    
    Returns:
        win32com.client.CDispatch: The Word application object.
    """
    try:
        word_app = win32com.client.Dispatch('Word.Application')
        word_app.Visible = False  # Set to True to make Word visible for debugging
        word_app.DisplayAlerts = False
        word_app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        return word_app
    except Exception as e:
        print(f"Error initializing Word application: {e}")
        exit(1)

word = initialize_word()

# Constants for Word
wdStatisticPages = 2

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Document Data"

# Write the header row
header = ['Document', 'Directory', 'N# Pages', 'N# Paragraphs', 'Size', 'Objective', 'Scope', 'Content']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20, 20, 50, 50, 50]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\Downloads'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

@retry_on_COM_error(max_retries=5, delay=2)
def get_page_count(doc):
    """
    Retrieves the number of pages in the Word document.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The number of pages.
    """
    return doc.ComputeStatistics(wdStatisticPages)

@retry_on_COM_error(max_retries=5, delay=2)
def get_paragraph_count(doc):
    """
    Retrieves the number of paragraphs in the Word document.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The number of paragraphs.
    """
    return doc.Paragraphs.Count

def get_toc_end_position(doc):
    """
    Returns the end position of the table of contents if it exists, otherwise returns 0.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The end position of the TOC or 0 if no TOC is found.
    """
    try:
        if doc.TablesOfContents.Count > 0:
            toc = doc.TablesOfContents(1)
            toc_range = toc.Range
            print(f"Table of contents found in '{doc.Name}'")
            return toc_range.End
    except Exception as e:
        print(f"Error identifying TOC in '{doc.Name}': {e}")
    return 0  # If no TOC is found, start from the beginning

def find_heading_position(doc, heading, heading_styles_set, start_pos, end_pos):
    """
    Searches for the specified heading within the document range and returns the end position of the heading if found.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading (str): The heading text to search for (case-insensitive).
        heading_styles_set (set): Set of normalized heading styles.
        start_pos (int): The starting position for the search.
        end_pos (int): The ending position for the search.

    Returns:
        int or None: The end position of the found heading. Returns None if not found.
    """
    current_pos = start_pos
    heading_lower = heading.lower()

    while current_pos < end_pos:
        para_range = doc.Range(Start=current_pos, End=end_pos)
        if para_range.Start >= para_range.End:
            break  # No more content
        para = para_range.Paragraphs(1)
        para_text = para.Range.Text.strip()
        para_style_name = para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in para_style_name.split(',')]

        # Debug
        # print(f"Searching for heading '{heading}': Found paragraph '{para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Remove numbering from the heading text
            heading_text_no_number = re.sub(r'^\d+(\.\d+)*\s*', '', para_text).strip()
            heading_text_no_number_lower = heading_text_no_number.lower()

            # Compare with the target heading
            if heading_lower == heading_text_no_number_lower:
                print(f"Matched heading '{heading_text_no_number}' at position {para.Range.End}")
                return para.Range.End  # Return the end position of the heading

        # Move to the next paragraph
        current_pos = para.Range.End

    # Debug
    # print(f"Heading '{heading}' not found between positions {start_pos} and {end_pos}.")
    return None  # Heading not found

def extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_pos):
    """
    Extracts content under a specific heading until any new heading is encountered.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading_end_pos (int): The position right after the heading.
        heading_styles_set (set): Set of normalized heading styles.
        end_pos (int): The end position of the search range.

    Returns:
        str: The extracted content under the heading.
    """
    content_paras = []
    current_content_pos = heading_end_pos

    while current_content_pos < end_pos:
        content_para_range = doc.Range(Start=current_content_pos, End=end_pos)
        if content_para_range.Start >= content_para_range.End:
            break  # No more content

        content_para = content_para_range.Paragraphs(1)
        content_para_text = content_para.Range.Text.strip()
        content_para_style_name = content_para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in content_para_style_name.split(',')]

        # Debug
        print(f"Extracting content: '{content_para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Reached any new heading, stop extraction
            print(f"Reached a new heading '{content_para_text}' with style variants {style_variants}. Stopping content extraction.")
            break

        # Add paragraph text to content if it's not empty
        if content_para_text:
            content_paras.append(content_para.Range.Text)

        # Move to the next paragraph
        current_content_pos = content_para.Range.End

    extracted_content = ''.join(content_paras).strip()
    print(f"Extracted content under heading ending at position {heading_end_pos}: {extracted_content[:60]}...")  # Show first 60 chars for brevity
    return extracted_content

def extract_section_content(doc, heading_texts, heading_styles_set, file_path):
    """
    Extracts and concatenates content under all specified headings within heading_texts, in the order they are listed.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading_texts (list): List of heading texts to search for.
        heading_styles_set (set): Set of normalized heading styles.
        file_path (str): Path to the Word document (used for debugging).

    Returns:
        str: The concatenated content of all matched headings.
    """
    content = ''
    MAX_PAGE = 13
    print(f"Searching limit is set to {MAX_PAGE} pages")
    try:
        print(f"Entering 'extract_section_content' for headings: {heading_texts}")

        # Get the end position of the TOC
        toc_end = get_toc_end_position(doc)
        print(f"TOC end position: {toc_end}")

        # Get the end position of page MAX_PAGE or the end of the document if less than MAX_PAGE pages
        try:
            page_max_page_range = doc.GoTo(What=1, Which=1, Count=MAX_PAGE)  # wdGoToPage=1, wdGoToAbsolute=1
            end_page_max_page = page_max_page_range.End
            print(f"End position of page {MAX_PAGE}: {end_page_max_page}")
        except Exception as e:
            print(f"Error getting end position of max_page {MAX_PAGE}: {e}")
            end_page_max_page = doc.Content.End

        # Check if end_page_max_page is before the TOC end
        if end_page_max_page < toc_end:
            print(f"End position of page {MAX_PAGE} is before the TOC end: {end_page_max_page}. Adjusting to document end.")
            end_page_max_page = doc.Content.End

        # Iterate through each heading in the specified order
        for heading in heading_texts:
            # Debug
            # print(f"Searching for heading '{heading}'")
            heading_end_pos = find_heading_position(doc, heading, heading_styles_set, toc_end, end_page_max_page)
            if heading_end_pos:
                # Extract content under this heading
                extracted_content = extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_page_max_page)
                if extracted_content:
                    print(f"Appending extracted content for heading '{heading}'")
                    content += extracted_content + "\n\n"  # Add double newline for separation
            else:
                print(f"Heading '{heading}' not found in document.")

        if content:
            print(f"Extracted content for headings {heading_texts}.")
        else:
            print(f"No content found for headings {heading_texts}.")

    except pywintypes.com_error as e:
        print(f"COM error in 'extract_section_content': {e}")
    except Exception as e:
        print(f"Error in 'extract_section_content': {e}")
    finally:
        print(f"Exiting 'extract_section_content' for headings: {heading_texts}")

    return content.strip()

# Define the heading styles to look for, ordered from highest to lowest level
heading_styles_ordered = [
    "Heading 1", "Heading 2", "Heading 3", "Heading 4", "Heading 5", 
    "Heading 6", "Heading 7", "Heading 8", "Titre 1", "Titre 2", 
    "Titre 3", "Titre 4", "Titre 5", "Titre 6", "Titre 7", "Titre 8"
]

# Convert heading styles to lowercase and create a set for faster lookup
heading_styles_set = set(style.lower() for style in heading_styles_ordered)

# Define the headings to search for each section
headings_dict = OrderedDict([
    ('objective', ['Objectif', 'But du document', 'Purpose', 'OBJECTIF ET CONTEXTE', 'Objet du document', "OBJET", "INTRODUCTION"]),
    ('scope', ['Périmètre', 'Périmètre fonctionnel']),
    ('content', ['Contenu', 'Content'])
])

# Main extraction loop
for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.doc', '.docx')):
            file_path = os.path.join(root, file)
            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            directory_path = os.path.dirname(file_path)

            # Increment the file counter before processing
            file_count += 1
            print(f"\n--------------------------------------------------------------------------------")
            print(f"Processing file {file_count}: {file_path}")
            print(f"File name: {file_name}")
            print(f"--------------------------------------------------------------------------------")

            try:
                print(f"Attempting to open document '{file_name}'")
                # Open the document in read-only mode
                doc = word.Documents.Open(file_path, ReadOnly=True)
                print(f"Successfully opened document '{file_name}'")
            except pywintypes.com_error as e:
                print(f"Error opening '{file_name}': {e}")
                skipped_files.append(file_path)
                continue
            except Exception as e:
                print(f"Unexpected error opening '{file_name}': {e}")
                skipped_files.append(file_path)
                continue

            size = ''
            pages = ''
            num_paragraphs = ''
            objective_content = ''
            scope_content = ''
            content_content = ''

            try:
                # Get the size of the document
                try:
                    size = os.path.getsize(file_path)
                    print(f"Size of document: {size} bytes")
                except Exception as e:
                    print(f"Error getting size: {e}")
                    size = ''

                # Get the total number of pages
                try:
                    pages = get_page_count(doc)
                    print(f"Total pages: {pages}")
                except Exception as e:
                    print(f"Error getting page count: {e}")
                    pages = ''

                # Get the number of paragraphs
                try:
                    num_paragraphs = get_paragraph_count(doc)
                    print(f"Number of paragraphs: {num_paragraphs}")
                except Exception as e:
                    print(f"Error getting paragraph count: {e}")
                    num_paragraphs = ''

                # Extract content for each section
                for section, heading_texts in headings_dict.items():
                    print(f"Extracting section '{section}' with headings: {heading_texts}")
                    content = extract_section_content(doc, heading_texts, heading_styles_set, file_path)
                    if section == 'objective':
                        objective_content = content
                    elif section == 'scope':
                        scope_content = content
                    elif section == 'content':
                        content_content = content

                # Sanitize the extracted content
                objective_content = remove_illegal_characters(objective_content.strip())
                scope_content = remove_illegal_characters(scope_content.strip())
                content_content = remove_illegal_characters(content_content.strip())
                file_name = remove_illegal_characters(file_name)
                file_path_sanitized = remove_illegal_characters(file_path)  # Avoid overwriting the original variable

                # Write the data to Excel
                ws.append([
                    file_name,
                    directory_path,
                    pages,
                    num_paragraphs,
                    size,
                    objective_content,
                    scope_content,
                    content_content
                ])
                print(f"Data written to Excel for '{file_name}'.")
            except Exception as e:
                print(f"Error processing '{file_name}': {e}")
                skipped_files.append(file_path)
            finally:
                # Ensure the document is closed
                try:
                    if 'doc' in locals() and doc is not None:
                        print(f"Attempting to close document '{file_name}'")
                        doc.Close(False)  # Close without saving
                        del doc
                        gc.collect()
                        print(f"Successfully closed document '{file_name}'")
                except Exception as e:
                    print(f"Error closing document '{file_name}': {e}")
                    word = initialize_word()

print(f"\nTotal files processed: {file_count}")

# Close the Word application
try:
    word.Quit()
    word = None  # Release the COM object
    gc.collect()
    print("Word application closed.")
except Exception as e:
    print(f"Error quitting Word application: {e}")

# Save the Excel workbook
try:
    wb.save('doc_output.xlsx')
    print("Excel workbook 'doc_output.xlsx' saved successfully.")
except Exception as e:
    print(f"Error saving Excel workbook: {e}")

# Display the skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files were processed successfully.")
