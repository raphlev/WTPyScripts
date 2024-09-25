import os
import re
import time
import gc
import functools
import pywintypes
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def remove_illegal_characters(value):
    if isinstance(value, str):
        # Remove illegal characters (control characters except for allowed ones)
        value = re.sub(
            r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]",
            "",
            value
        )
    return value

# Retry decorator for COM calls with enhanced logging
def retry_on_COM_error(max_retries=3, delay=1):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            while retries <= max_retries:
                try:
                    return func(*args, **kwargs)
                except pywintypes.com_error as e:
                    if e.args[0] == -2147418111:  # Call was rejected by callee
                        retries += 1
                        print(f"COM call to '{func.__name__}' failed with 'Call was rejected by callee', retrying ({retries}/{max_retries})...")
                        time.sleep(delay)
                    else:
                        print(f"COM error in '{func.__name__}': {e}")
                        raise
                except Exception as e:
                    print(f"Unexpected error in '{func.__name__}': {e}")
                    raise
            print(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
            raise Exception(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
        return wrapper
    return decorator

# Initialize the Word application
def initialize_word():
    try:
        word_app = win32com.client.Dispatch('Word.Application')
        word_app.Visible = True  # Set to True to make Word visible for debugging
        word_app.DisplayAlerts = False
        word_app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        return word_app
    except Exception as e:
        print(f"Error initializing Word application: {e}")
        exit(1)

word = initialize_word()

# Constants for Word
wdStatisticPages = 2

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Document Data"

# Write the header row
header = ['Document FileName', 'Document FilePath', 'Total number of pages', 'Number of Paragraphs', 'Size of document', 'Objective', 'Scope', 'Content']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20, 20, 50, 50, 50]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\DOCUMENTATION_APPLICATIONS'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

@retry_on_COM_error(max_retries=5, delay=2)
def get_page_count(doc):
    return doc.ComputeStatistics(wdStatisticPages)

@retry_on_COM_error(max_retries=5, delay=2)
def get_paragraph_count(doc):
    return doc.Paragraphs.Count

@retry_on_COM_error(max_retries=5, delay=2)
def get_document_content(doc):
    return doc.Content

def extract_section_content(doc, heading_texts, found_headings, file_path):
    """
    Extracts content under one or more headings until the next heading of the same style.
    """
    content = ''
    try:
        print(f"Entering 'extract_section_content' for file '{file_path}', headings '{heading_texts}'")
        rng = get_document_content(doc)
        for heading_text in heading_texts:
            if heading_text.lower() in found_headings:
                print(f"Heading '{heading_text}' already found, skipping.")
                continue  # Skip if this heading has already been processed

            # Prepare the Find object
            rng.Find.ClearFormatting()
            rng.Find.Text = heading_text
            rng.Find.MatchCase = False
            rng.Find.MatchWholeWord = True
            rng.Find.Forward = True
            rng.Find.Format = False  # Set to False since we're not using formatting here

            print(f"Searching for heading '{heading_text}' in '{file_path}'")
            found = rng.Find.Execute()
            if found:
                print(f"Found heading '{heading_text}' in '{file_path}'")
                found_headings.add(heading_text.lower())

                start_range = rng.Duplicate
                start_range.Collapse(0)  # Move after the heading

                end_range = get_document_content(doc)
                end_range.Start = start_range.Start

                heading_found = False

                # Find the next heading of the same style
                for style_name in heading_style_names:
                    try:
                        end_range.Find.ClearFormatting()
                        end_range.Find.Style = doc.Styles(style_name)
                    except Exception as e:
                        print(f"Style '{style_name}' not found in '{file_path}': {e}")
                        continue  # Skip if the style is not found

                    end_range.Find.Text = ""
                    end_range.Find.MatchWildcards = True
                    end_range.Find.MatchCase = False
                    end_range.Find.MatchWholeWord = False
                    end_range.Find.Forward = True
                    end_range.Find.Format = True

                    print(f"Searching for next heading of style '{style_name}' in '{file_path}'")
                    end_found = end_range.Find.Execute(FindText="*")
                    if end_found:
                        print(f"Found next heading of style '{style_name}' in '{file_path}'")
                        end_range.Collapse(1)  # Collapse to start
                        start_range.End = end_range.Start
                        heading_found = True
                        break

                if not heading_found:
                    print(f"No next heading found in '{file_path}'")
                    start_range.End = doc.Content.End

                content = start_range.Text.strip()
                print(f"Extracted content for heading '{heading_text}' in '{file_path}'.")
                break  # Exit after finding the first matching heading
            else:
                print(f"Heading '{heading_text}' not found in '{file_path}'")
    except pywintypes.com_error as e:
        print(f"COM error in 'extract_section_content' for '{file_path}': {e}")
    except Exception as e:
        print(f"Error in 'extract_section_content' for '{file_path}': {e}")
    finally:
        print(f"Exiting 'extract_section_content' for file '{file_path}', headings '{heading_texts}'")
    return content

# Define the heading styles to look for
heading_style_names = ["Heading 1", "Heading 2", "Titre 1", "Titre 2"]

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.doc', '.docx')):
            file_path = os.path.join(root, file)

            # Increment the file counter before processing
            file_count += 1

            print(f"\nProcessing file {file_count}: {file_path}")
            print(f"File name: {file}")

            try:
                print(f"Attempting to open document '{file_path}'")
                # Open the document in read-only mode
                doc = word.Documents.Open(file_path, ReadOnly=True)
                print(f"Successfully opened document '{file_path}'")
            except pywintypes.com_error as e:
                print(f"Error opening '{file_path}': {e}")
                skipped_files.append(file_path)
                continue
            except Exception as e:
                print(f"Unexpected error opening '{file_path}': {e}")
                skipped_files.append(file_path)
                continue

            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            size = ''
            pages = ''
            num_paragraphs = ''
            objective_content = ''
            scope_content = ''
            content_content = ''

            try:
                # Get the size of the document
                try:
                    size = os.path.getsize(file_path)
                    print(f"Size of '{file_path}': {size} bytes")
                except Exception as e:
                    print(f"Error getting size for '{file_path}': {e}")
                    size = ''

                # Get the total number of pages
                try:
                    pages = get_page_count(doc)
                    print(f"Total pages in '{file_path}': {pages}")
                except Exception as e:
                    print(f"Error getting page count for '{file_path}': {e}")
                    pages = ''

                # Get the number of paragraphs
                try:
                    num_paragraphs = get_paragraph_count(doc)
                    print(f"Number of paragraphs in '{file_path}': {num_paragraphs}")
                except Exception as e:
                    print(f"Error getting paragraph count for '{file_path}': {e}")
                    num_paragraphs = ''

                # Set to keep track of already found headings
                found_headings = set()

                # Define the headings to search for each section
                headings_dict = {
                    'objective': ['Objectif', 'But du document', 'Purpose', 'OBJECTIF ET CONTEXTE'],
                    'scope': ['Périmètre'],
                    'content': ['Contenu', 'Content']
                }

                # Extract content for each section
                for section, heading_texts in headings_dict.items():
                    print(f"Extracting section '{section}' for '{file_path}'")
                    content = extract_section_content(doc, heading_texts, found_headings, file_path)
                    if section == 'objective':
                        objective_content = content
                    elif section == 'scope':
                        scope_content = content
                    elif section == 'content':
                        content_content = content

                # Sanitize the extracted content
                objective_content = remove_illegal_characters(objective_content.strip())
                scope_content = remove_illegal_characters(scope_content.strip())
                content_content = remove_illegal_characters(content_content.strip())
                file_name = remove_illegal_characters(file_name)
                file_path = remove_illegal_characters(file_path)

                # Write the data to Excel
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    num_paragraphs,
                    size,
                    objective_content,
                    scope_content,
                    content_content
                ])
                print(f"Data written to Excel for '{file_name}'.")
            except Exception as e:
                print(f"Error processing '{file_path}': {e}")
                skipped_files.append(file_path)
            finally:
                # Ensure the document is closed
                try:
                    if doc is not None:
                        print(f"Attempting to close document '{file_path}'")
                        doc.Close(False)  # Close without saving
                        del doc
                        gc.collect()
                        print(f"Successfully closed document '{file_path}'")
                except Exception as e:
                    print(f"Error closing document '{file_path}': {e}")
                    word = initialize_word()

print(f"\nTotal files processed: {file_count}")

# Close the Word application
try:
    word.Quit()
    word = None  # Release the COM object
    gc.collect()
    print("Word application closed.")
except Exception as e:
    print(f"Error quitting Word application: {e}")

# Save the Excel workbook
try:
    wb.save('output.xlsx')
    print("Excel workbook 'output.xlsx' saved successfully.")
except Exception as e:
    print(f"Error saving Excel workbook: {e}")

# Display the skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files were processed successfully.")
