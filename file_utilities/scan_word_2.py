import os
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Ensure you have openpyxl installed:
# pip install openpyxl

# Initialize Word application
try:
    word = win32com.client.Dispatch('Word.Application')
    word.Visible = False  # False - Set to True temporarily for debugging
    word.DisplayAlerts = False # Suppress Alerts in Word

except Exception as e:
    print(f"Error initializing Word application: {e}")
    exit(1)

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Document Data"

# Write the header row
header = ['Document FileName', 'Document FilePath', 'Total number of pages', 'Size of document', 'Objectif', 'Périmètre', 'Contenu']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20, 50, 50, 50]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\DOCUMENTATION_APPLICATIONS\Opale'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.doc', '.docx')):
            file_path = os.path.join(root, file)
            print(f"Processing file {file_count+1}: {file_path}")  # Progress message
            try:
                # Open the document in ReadOnly mode
                doc = word.Documents.Open(file_path, ReadOnly=True)
            except Exception as e:
                print(f"Error opening {file_path}: {e}")
                skipped_files.append(file_path)
                continue

            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            size = ''
            pages = ''
            objectif_content = ''
            perimetre_content = ''
            contenu_content = ''

            try:
                # Get Size of document
                try:
                    size = os.path.getsize(file_path)
                except Exception as e:
                    print(f"Error getting size for {file_path}: {e}")
                    size = ''

                # Get Total number of pages
                try:
                    pages = doc.ComputeStatistics(2)  # 2 corresponds to wdStatisticPages
                except Exception as e:
                    print(f"Error getting page count for {file_path}: {e}")
                    pages = ''

                # Variables to control collection
                collecting = None

                # Total number of paragraphs
                num_paragraphs = doc.Paragraphs.Count
                print(f"Total paragraphs in document: {num_paragraphs}")

                # Iterate over paragraphs by index
                for i in range(1, num_paragraphs + 1):
                    try:
                        para = doc.Paragraphs(i)
                        style = para.Style.NameLocal.lower()
                        text = para.Range.Text.strip()
                    except Exception as e:
                        # print(f"Error accessing paragraph {i} in {file_path}: {e}")
                        continue  # Skip this paragraph

                    # Check if paragraph is a heading
                    if 'heading' in style or 'titre' in style:
                        # This is a heading
                        heading_text = text.lower()
                        if heading_text == 'objectif':
                            collecting = 'objectif'
                            continue  # Skip the heading text itself
                        elif heading_text == 'périmètre':
                            collecting = 'perimetre'
                            continue
                        elif heading_text == 'contenu':
                            collecting = 'contenu'
                            continue
                        else:
                            # New heading, stop collecting
                            collecting = None
                    else:
                        if collecting == 'objectif':
                            objectif_content += text + '\n'
                        elif collecting == 'perimetre':
                            perimetre_content += text + '\n'
                        elif collecting == 'contenu':
                            contenu_content += text + '\n'

                # Write data to Excel
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])

                file_count += 1  # Increment processed file count

            except Exception as e:
                print(f"Error processing {file_path}: {e}")
                # Write collected data even if there was an error
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])
            finally:
                # Ensure the document is closed
                try:
                    if doc is not None:
                        doc.Close(False)  # Close without saving
                except Exception as e:
                    print(f"Error closing document {file_path}: {e}")
                finally:
                    doc = None

print(f"\nTotal files processed: {file_count}")

# Close Word application
word.Quit()
word = None  # Release the COM object

# Save the Excel workbook
wb.save('output.xlsx')

# Print skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files processed successfully.")
