import os
import re
import time
import gc
import functools
import pywintypes
import win32com.client
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging
import pythoncom
from multiprocessing import Pool, cpu_count

# Configure logging
logging.basicConfig(
    filename='doc_processing.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

def remove_illegal_characters(value):
    """
    Removes illegal control characters from a string, except for allowed ones.
    
    Args:
        value (str): The string to sanitize.
    
    Returns:
        str: The sanitized string.
    """
    if isinstance(value, str):
        # Remove illegal characters (control characters except for allowed ones)
        value = illegal_char_pattern.sub("", value)
    return value

# Compile the regular expression pattern for removing illegal characters
illegal_char_pattern = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]")

# Compile the regular expression pattern for removing numbering from headings
heading_number_pattern = re.compile(r'^\d+(\.\d+)*\s*')

# Retry decorator for COM calls with enhanced logging
def retry_on_COM_error(max_retries=3, delay=1):
    """
    Decorator to retry a function upon encountering a COM error.
    
    Args:
        max_retries (int): Maximum number of retries.
        delay (int): Delay in seconds between retries.
    
    Returns:
        function: The decorated function with retry logic.
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            while retries <= max_retries:
                try:
                    return func(*args, **kwargs)
                except pywintypes.com_error as e:
                    if e.args[0] == -2147418111:  # Call was rejected by callee
                        retries += 1
                        logging.warning(f"COM call to '{func.__name__}' failed with 'Call was rejected by callee', retrying ({retries}/{max_retries})...")
                        time.sleep(delay)
                    else:
                        logging.error(f"COM error in '{func.__name__}': {e}")
                        raise
                except Exception as e:
                    logging.error(f"Unexpected error in '{func.__name__}': {e}")
                    raise
            logging.critical(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
            raise Exception(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
        return wrapper
    return decorator

# Initialize the Word application using early binding
def initialize_word():
    """
    Initializes and returns the Word application object using early binding.
    
    Returns:
        win32com.client.CDispatch: The Word application object.
    """
    try:
        pythoncom.CoInitialize()
        word_app = win32com.client.gencache.EnsureDispatch('Word.Application')
        word_app.Visible = False  # Set to True to make Word visible for debugging
        word_app.DisplayAlerts = False
        word_app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        word_app.AskToUpdateLinks = False  # Prevent prompts to update links
        return word_app
    except Exception as e:
        logging.critical(f"Error initializing Word application: {e}")
        exit(1)

word = initialize_word()

# Constants for Word
wdStatisticPages = 2

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Document Data"

# Write the header row
header = ['Document', 'Directory', 'N# Pages', 'N# Paragraphs', 'Size', 'Objective', 'Scope', 'Content']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20, 20, 50, 50, 50]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\DOCUMENTATION_APPLICATIONS'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# List to accumulate data for Excel
extracted_data_list = []

# Counter for processed files
file_count = 0

@retry_on_COM_error(max_retries=5, delay=2)
def get_page_count(doc):
    """
    Retrieves the number of pages in the Word document.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The number of pages.
    """
    return doc.ComputeStatistics(wdStatisticPages)

@retry_on_COM_error(max_retries=5, delay=2)
def get_paragraph_count(doc):
    """
    Retrieves the number of paragraphs in the Word document.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The number of paragraphs.
    """
    return doc.Paragraphs.Count

def get_toc_end_position(doc):
    """
    Returns the end position of the table of contents if it exists, otherwise returns 0.
    
    Args:
        doc (win32com.client.CDispatch): The Word document object.
    
    Returns:
        int: The end position of the TOC or 0 if no TOC is found.
    """
    try:
        if doc.TablesOfContents.Count > 0:
            toc = doc.TablesOfContents(1)
            toc_range = toc.Range
            logging.info(f"Table of contents found in '{doc.Name}'")
            return toc_range.End
    except Exception as e:
        logging.error(f"Error identifying TOC in '{doc.Name}': {e}")
    return 0  # If no TOC is found, start from the beginning

def find_heading_position(doc, heading, heading_styles_set, start_pos, end_pos, heading_number_pattern):
    """
    Searches for the specified heading within the document range and returns the end position of the heading if found.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading (str): The heading text to search for (case-insensitive).
        heading_styles_set (set): Set of normalized heading styles.
        start_pos (int): The starting position for the search.
        end_pos (int): The ending position for the search.
        heading_number_pattern (re.Pattern): Compiled regex pattern for removing numbering.

    Returns:
        int or None: The end position of the found heading. Returns None if not found.
    """
    current_pos = start_pos
    heading_lower = heading.lower()

    while current_pos < end_pos:
        para_range = doc.Range(Start=current_pos, End=end_pos)
        if para_range.Start >= para_range.End:
            break  # No more content
        para = para_range.Paragraphs(1)
        para_text = para.Range.Text.strip()
        para_style_name = para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in para_style_name.split(',')]

        # Debug
        logging.debug(f"Searching for heading '{heading}': Found paragraph '{para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Remove numbering from the heading text using the compiled pattern
            heading_text_no_number = heading_number_pattern.sub('', para_text).strip()
            heading_text_no_number_lower = heading_text_no_number.lower()

            # Compare with the target heading
            if heading_lower == heading_text_no_number_lower:
                logging.info(f"Matched heading '{heading_text_no_number}' at position {para.Range.End}")
                return para.Range.End  # Return the end position of the heading

        # Move to the next paragraph
        current_pos = para.Range.End

    logging.warning(f"Heading '{heading}' not found between positions {start_pos} and {end_pos}.")
    return None  # Heading not found

def extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_pos, heading_number_pattern):
    """
    Extracts content under a specific heading until any new heading is encountered.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading_end_pos (int): The position right after the heading.
        heading_styles_set (set): Set of normalized heading styles.
        end_pos (int): The end position of the search range.
        heading_number_pattern (re.Pattern): Compiled regex pattern for removing numbering.

    Returns:
        str: The extracted content under the heading.
    """
    content_paras = []
    current_content_pos = heading_end_pos

    while current_content_pos < end_pos:
        content_para_range = doc.Range(Start=current_content_pos, End=end_pos)
        if content_para_range.Start >= content_para_range.End:
            break  # No more content

        content_para = content_para_range.Paragraphs(1)
        content_para_text = content_para.Range.Text.strip()
        content_para_style_name = content_para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in content_para_style_name.split(',')]

        # Debug
        logging.debug(f"Extracting content: '{content_para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Reached any new heading, stop extraction
            logging.info(f"Reached a new heading '{content_para_text}' with style variants {style_variants}. Stopping content extraction.")
            break

        # Add paragraph text to content if it's not empty
        if content_para_text:
            content_paras.append(content_para.Range.Text)

        # Move to the next paragraph
        current_content_pos = content_para.Range.End

    extracted_content = ''.join(content_paras).strip()
    logging.info(f"Extracted content under heading ending at position {heading_end_pos}: {extracted_content[:60]}...")  # Show first 60 chars for brevity
    return extracted_content

def extract_section_content(doc, heading_texts, heading_styles_set, file_path, heading_number_pattern):
    """
    Extracts and concatenates content under all specified headings within heading_texts, in the order they are listed.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading_texts (list): List of heading texts to search for.
        heading_styles_set (set): Set of normalized heading styles.
        file_path (str): Path to the Word document (used for debugging).
        heading_number_pattern (re.Pattern): Compiled regex pattern for removing numbering.

    Returns:
        str: The concatenated content of all matched headings.
    """
    content = ''
    MAX_PAGE = 13
    print(f"Searching limit is set to {MAX_PAGE} pages")
    try:
        logging.info(f"Entering 'extract_section_content' for headings: {heading_texts}")

        # Get the end position of the TOC
        toc_end = get_toc_end_position(doc)
        logging.info(f"TOC end position: {toc_end}")

        # Get the end position of page MAX_PAGE or the end of the document if less than MAX_PAGE pages
        try:
            page_max_page_range = doc.GoTo(What=1, Which=1, Count=MAX_PAGE)  # wdGoToPage=1, wdGoToAbsolute=1
            end_page_max_page = page_max_page_range.End
            logging.info(f"End position of page {MAX_PAGE}: {end_page_max_page}")
        except Exception as e:
            logging.error(f"Error getting end position of max_page {MAX_PAGE}: {e}")
            end_page_max_page = doc.Content.End

        # Check if end_page_max_page is before the TOC end
        if end_page_max_page < toc_end:
            logging.info(f"End position of page {MAX_PAGE} is before the TOC end: {end_page_max_page}. Adjusting to document end.")
            end_page_max_page = doc.Content.End

        # Iterate through each heading in the specified order
        for heading in heading_texts:
            logging.info(f"Searching for heading '{heading}'")
            heading_end_pos = find_heading_position(doc, heading, heading_styles_set, toc_end, end_page_max_page, heading_number_pattern)
            if heading_end_pos:
                # Extract content under this heading
                extracted_content = extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_page_max_page, heading_number_pattern)
                if extracted_content:
                    logging.info(f"Appending extracted content for heading '{heading}'")
                    content += extracted_content + "\n\n"  # Add double newline for separation
            else:
                logging.warning(f"Heading '{heading}' not found in document.")

        if content:
            logging.info(f"Extracted content for headings {heading_texts}.")
        else:
            logging.info(f"No content found for headings {heading_texts}.")

    except pywintypes.com_error as e:
        logging.error(f"COM error in 'extract_section_content': {e}")
    except Exception as e:
        logging.error(f"Error in 'extract_section_content': {e}")
    finally:
        logging.info(f"Exiting 'extract_section_content' for headings: {heading_texts}")

    return content.strip()

# Define the heading styles to look for, ordered from highest to lowest level
heading_styles_ordered = [
    "Heading 1", "Heading 2", "Heading 3", "Heading 4", "Heading 5", 
    "Heading 6", "Heading 7", "Heading 8", "Titre 1", "Titre 2", 
    "Titre 3", "Titre 4", "Titre 5", "Titre 6", "Titre 7", "Titre 8"
]

# Convert heading styles to lowercase and create a set for faster lookup
heading_styles_set = set(style.lower() for style in heading_styles_ordered)

# Define the headings to search for each section
headings_dict = OrderedDict([
    ('objective', ['Objectif', 'But du document', 'Purpose', 'OBJECTIF ET CONTEXTE', 'Objet du document', "OBJET", "INTRODUCTION"]),
    ('scope', ['Périmètre', 'Périmètre fonctionnel']),
    ('content', ['Contenu', 'Content'])
])

# Main extraction loop
def main():
    global file_count  # To modify the global file_count variable
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if file.lower().endswith(('.doc', '.docx')):
                file_path = os.path.join(root, file)
                # Initialize variables with default empty values
                file_name = os.path.basename(file_path)
                directory_path = os.path.dirname(file_path)

                # Increment the file counter before processing
                file_count += 1
                logging.info(f"\n--------------------------------------------------------------------------------")
                logging.info(f"Processing file {file_count}: {file_path}")
                logging.info(f"File name: {file_name}")
                logging.info(f"--------------------------------------------------------------------------------")

                try:
                    logging.info(f"Attempting to open document '{file_name}'")
                    # Open the document in read-only mode without updating links
                    doc = word.Documents.Open(
                        FileName=file_path,
                        ConfirmConversions=False,
                        ReadOnly=True,
                        AddToRecentFiles=False,
                        PasswordDocument=None,
                        UpdateLinks=0  # 0 = wdUpdateLinksNever
                    )
                    logging.info(f"Successfully opened document '{file_name}'")
                except pywintypes.com_error as e:
                    logging.error(f"Error opening '{file_name}': {e}")
                    skipped_files.append(file_path)
                    continue
                except Exception as e:
                    logging.error(f"Unexpected error opening '{file_name}': {e}")
                    skipped_files.append(file_path)
                    continue

                size = ''
                pages = ''
                num_paragraphs = ''
                objective_content = ''
                scope_content = ''
                content_content = ''

                try:
                    # Get the size of the document
                    try:
                        size = os.path.getsize(file_path)
                        logging.info(f"Size of document: {size} bytes")
                    except Exception as e:
                        logging.error(f"Error getting size: {e}")
                        size = ''

                    # Get the total number of pages
                    try:
                        pages = get_page_count(doc)
                        logging.info(f"Total pages: {pages}")
                    except Exception as e:
                        logging.error(f"Error getting page count: {e}")
                        pages = ''

                    # Get the number of paragraphs
                    try:
                        num_paragraphs = get_paragraph_count(doc)
                        logging.info(f"Number of paragraphs: {num_paragraphs}")
                    except Exception as e:
                        logging.error(f"Error getting paragraph count: {e}")
                        num_paragraphs = ''

                    # Extract content for each section
                    for section, heading_texts in headings_dict.items():
                        logging.info(f"Extracting section '{section}' with headings: {heading_texts}")
                        content = extract_section_content(doc, heading_texts, heading_styles_set, file_path, heading_number_pattern)
                        if section == 'objective':
                            objective_content = content
                        elif section == 'scope':
                            scope_content = content
                        elif section == 'content':
                            content_content = content

                    # Sanitize the extracted content
                    objective_content = remove_illegal_characters(objective_content.strip())
                    scope_content = remove_illegal_characters(scope_content.strip())
                    content_content = remove_illegal_characters(content_content.strip())
                    file_name_sanitized = remove_illegal_characters(file_name)
                    file_path_sanitized = remove_illegal_characters(file_path)  # Avoid overwriting the original variable

                    # Append the data to the accumulation list
                    extracted_data_list.append([
                        file_name_sanitized,
                        directory_path,
                        pages,
                        num_paragraphs,
                        size,
                        objective_content,
                        scope_content,
                        content_content
                    ])
                    logging.info(f"Data appended to memory for '{file_name}'.")
                except Exception as e:
                    logging.error(f"Error processing '{file_name}': {e}")
                    skipped_files.append(file_path)
                finally:
                    # Ensure the document is closed and COM object is released
                    try:
                        if 'doc' in locals() and doc is not None:
                            logging.info(f"Attempting to close document '{file_name}'")
                            doc.Close(False)  # Close without saving
                            del doc
                            gc.collect()
                            logging.info(f"Successfully closed document '{file_name}'")
                    except Exception as e:
                        logging.error(f"Error closing document '{file_name}': {e}")
                        # Re-initialize Word to recover from errors
                        global word
                        word = initialize_word()

    # Start the main processing
    main()

    logging.info(f"\nTotal files processed: {file_count}")

    # Write all accumulated data to Excel at once
    try:
        logging.info("Starting to write accumulated data to Excel.")
        for data in extracted_data_list:
            ws.append(data)
        logging.info("All accumulated data has been written to Excel.")
    except Exception as e:
        logging.error(f"Error writing data to Excel: {e}")

    # Close the Word application
    try:
        word.Quit()
        word = None  # Release the COM object
        gc.collect()
        logging.info("Word application closed.")
    except Exception as e:
        logging.error(f"Error quitting Word application: {e}")

    # Save the Excel workbook
    try:
        wb.save('doc_output.xlsx')
        logging.info("Excel workbook 'doc_output.xlsx' saved successfully.")
    except Exception as e:
        logging.error(f"Error saving Excel workbook: {e}")

    # Display the skipped files
    if skipped_files:
        logging.warning("\nThe following files were skipped due to errors:")
        for f in skipped_files:
            logging.warning(f)
    else:
        logging.info("\nAll files were processed successfully.")

# Execute the script
if __name__ == '__main__':
    main()
