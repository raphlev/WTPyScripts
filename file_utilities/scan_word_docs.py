import os
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Assurez-vous d'installer openpyxl si vous ne l'avez pas déjà :
# pip install openpyxl

# Initialisation de l'application Word
try:
    word = win32com.client.Dispatch('Word.Application')
    word.Visible = False
except Exception as e:
    print(f"Erreur lors de l'initialisation de l'application Word : {e}")
    exit(1)

# Création d'un nouveau classeur Excel et sélection de la feuille active
wb = Workbook()
ws = wb.active
ws.title = "Données des documents"

# Écriture de la ligne d'en-tête
header = ['Document FileName', 'Document FilePath', 'Total number of pages', 'Size of document', 'Objectif', 'Périmètre', 'Contenu']
ws.append(header)

# Ajustement des largeurs de colonnes (optionnel)
column_widths = [30, 100, 20, 20, 50, 50, 50]  # Largeurs pour chaque colonne
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Répertoire racine à scanner
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\DOCUMENTATION_APPLICATIONS\Indigo'  # Remplacez par le chemin de votre dossier

# Liste pour garder une trace des fichiers ignorés
skipped_files = []

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.doc', '.docx')):
            file_path = os.path.join(root, file)
            try:
                # Ouverture du document en mode lecture seule
                doc = word.Documents.Open(file_path, ReadOnly=True)
            except Exception as e:
                print(f"Erreur lors de l'ouverture de {file_path} : {e}")
                skipped_files.append(file_path)
                continue

            # Initialisation des variables avec des valeurs vides par défaut
            file_name = os.path.basename(file_path)
            size = ''
            pages = ''
            objectif_content = ''
            perimetre_content = ''
            contenu_content = ''

            try:
                # Obtention de la taille du document
                try:
                    size = os.path.getsize(file_path)
                except Exception as e:
                    print(f"Erreur lors de l'obtention de la taille pour {file_path} : {e}")
                    size = ''

                # Obtention du nombre total de pages
                try:
                    pages = doc.ComputeStatistics(2)  # 2 correspond à wdStatisticPages
                except Exception as e:
                    print(f"Erreur lors de l'obtention du nombre de pages pour {file_path} : {e}")
                    pages = ''

                # Variables pour contrôler la collecte
                collecting = None

                # Nombre total de paragraphes
                num_paragraphs = doc.Paragraphs.Count

                # Parcours des paragraphes par index
                for i in range(1, num_paragraphs + 1):
                    try:
                        para = doc.Paragraphs(i)
                        style = para.Style.NameLocal.lower()
                        text = para.Range.Text.strip()
                    except Exception as e:
                        print(f"Erreur lors de l'accès au paragraphe {i} dans {file_path} : {e}")
                        continue  # Passer ce paragraphe

                    # Vérifier si le paragraphe est un titre
                    if 'heading' in style or 'titre' in style:
                        # C'est un titre
                        heading_text = text.lower()
                        if heading_text == 'objectif':
                            collecting = 'objectif'
                            continue  # Passer le texte du titre lui-même
                        elif heading_text == 'périmètre':
                            collecting = 'perimetre'
                            continue
                        elif heading_text == 'contenu':
                            collecting = 'contenu'
                            continue
                        else:
                            # Nouveau titre, arrêter la collecte
                            collecting = None
                    else:
                        if collecting == 'objectif':
                            objectif_content += text + '\n'
                        elif collecting == 'perimetre':
                            perimetre_content += text + '\n'
                        elif collecting == 'contenu':
                            contenu_content += text + '\n'

                # Écriture des données dans Excel
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])

            except Exception as e:
                print(f"Erreur lors du traitement de {file_path} : {e}")
                # Écrire les données collectées même s'il y a une erreur
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])
            finally:
                # Fermer le document en s'assurant qu'il est toujours valide
                try:
                    if doc is not None:
                        doc.Close(False)  # Fermer sans enregistrer
                except Exception as e:
                    print(f"Erreur lors de la fermeture du document {file_path} : {e}")
                finally:
                    # Libérer l'objet COM du document
                    doc = None

# Fermer l'application Word
word.Quit()
word = None  # Libérer l'objet COM de Word

# Sauvegarder le classeur Excel
wb.save('output.xlsx')

# Afficher les fichiers ignorés
if skipped_files:
    print("\nLes fichiers suivants ont été ignorés en raison d'erreurs :")
    for f in skipped_files:
        print(f)
else:
    print("\nTous les fichiers ont été traités avec succès.")
