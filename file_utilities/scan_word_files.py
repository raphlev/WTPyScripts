r"""
===============================================================================
Document Processing and Data Extraction Script
===============================================================================

**Author:** Your Name  
**Date:** 2024-09-27  
**Version:** 1.0.0

-----------------------------------------------------------------------------
**Description:**
-----------------------------------------------------------------------------
This Python script automates the process of scanning through a specified 
directory (and its subdirectories) to locate Microsoft Word documents 
(`.doc` and `.docx` files). For each document found, the script performs the 
following actions:

1. **Initialization:**
   - Launches a hidden instance of Microsoft Word using COM automation.
   - Configures Word application settings to disable macros for security.

2. **Data Extraction:**
   - Opens each Word document in read-only mode.
   - Computes the number of pages and paragraphs in the document.
   - Identifies the end position of the Table of Contents (TOC) if it exists.
   - Searches for specific headings (e.g., "Objective," "Scope," "Content") 
     within the document.
   - Extracts and sanitizes content under these headings, removing any 
     illegal control characters.

3. **Data Compilation:**
   - Aggregates the extracted information, including document name, directory path, 
     page count, paragraph count, file size, and the content under each specified section.

4. **Output Generation:**
   - Writes the compiled data into an Excel workbook (`doc_output.xlsx`) with 
     predefined headers and column widths for readability.
   - Maintains a log file (`doc_processing.log`) to record the script's activities, 
     warnings, and errors for debugging and auditing purposes.

5. **Cleanup:**
   - Ensures that the Word application is properly closed after processing to 
     prevent orphaned processes.

-----------------------------------------------------------------------------
**Dependencies:**
-----------------------------------------------------------------------------
- **Python 3.6+**
- **Libraries:**
  - `pywin32` (for COM automation with Microsoft Word)
  - `openpyxl` (for Excel file creation and manipulation)
  - `logging` (for logging script activities)
  - Standard Python libraries: `os`, `re`, `time`, `gc`, `functools`, `collections`, `pythoncom`, `atexit`

-----------------------------------------------------------------------------
**Installation:**
-----------------------------------------------------------------------------
Ensure that all dependencies are installed. You can install the required 
libraries using `pip`:

```bash
pip install pywin32 openpyxl
```

-----------------------------------------------------------------------------
**Usage:**
-----------------------------------------------------------------------------
- Configuration:
  Root Directory: Modify the root_dir variable in the main() function to specify the directory you want to scan for Word documents.
```
root_dir = r'C:\Path\To\Your\Documents'  # Replace with your target directory
```
  Logging Level: The logging level is set to INFO by default. For more detailed logs, you can change it to DEBUG in the logging configuration section.
```
logging.basicConfig(
    filename='doc_processing.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO  # Change to logging.DEBUG for detailed logs
)
```
  Heading Styles and Sections: Customize the headings_dict and heading_styles_ordered to match the specific headings and styles used in your Word documents.

- Execution:
  Direct Execution: Run the script using Python from the command line:
```
python your_script_name.py
```
  Batch File (Optional): You can create a .bat file to automate execution and monitor logs simultaneously. Refer to the assistant's previous messages for guidance on setting this up.

- Output:
  Excel Workbook: After execution, check the doc_output.xlsx file for the compiled data.
  Log File: Review the doc_processing.log file for detailed logs, including any errors or warnings encountered during processing.

-----------------------------------------------------------------------------
**Error Handling:**
-----------------------------------------------------------------------------
- The script includes robust error handling mechanisms:
  COM Errors: Utilizes a decorator (retry_on_COM_error) to retry functions that fail due to transient COM errors.
  Logging: All errors, warnings, and critical issues are logged in doc_processing.log for easy debugging.
  Cleanup: Ensures that the Word application is properly closed even if unexpected errors occur, preventing orphaned Word processes.

-----------------------------------------------------------------------------
**Customization:**
-----------------------------------------------------------------------------
- Headings: Modify the headings_dict to include all the headings you wish to extract from your documents.
- Output Fields: Adjust the headers and data fields in the setup_excel_workbook and process_document functions to capture additional or different information as needed.

===============================================================================
"""
import os
import re
import time
import gc
import functools
import pywintypes
import win32com.client
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging
import pythoncom
import atexit  # For cleanup upon exit

# ============================
# 1. Configuration and Setup
# ============================

# Configure logging with enhanced details
logging.basicConfig(
    filename='doc_processing.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO  # Set to DEBUG for detailed logs
)

# Initialize global variables
word = None  # Will hold the Word application object

# ============================
# 2. Cleanup Mechanism
# ============================

def cleanup_word():
    """
    Cleans up the Word application by quitting it if it's still running.
    This function is registered with atexit to ensure it's called upon script exit.
    """
    global word
    if word is not None:
        try:
            logging.info("Attempting to quit Word application during cleanup.")
            word.Quit()
            word = None
            logging.info("Word application closed successfully during cleanup.")
        except Exception as e:
            logging.error(f"Error quitting Word application during cleanup: {e}")

# Register the cleanup function to be called upon script exit
atexit.register(cleanup_word)

# ============================
# 3. Helper Functions
# ============================

def remove_illegal_characters(value):
    """
    Removes illegal control characters from a string, except for allowed ones.

    Args:
        value (str): The string to sanitize.

    Returns:
        str: The sanitized string.
    """
    if isinstance(value, str):
        # Remove illegal characters (control characters except for allowed ones)
        return illegal_char_pattern.sub("", value)
    return value

# Compile the regular expression pattern for removing illegal characters
illegal_char_pattern = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]")

# Compile the regular expression pattern for removing numbering from headings
heading_number_pattern = re.compile(r'^\d+(\.\d+)*\s*')

def retry_on_COM_error(max_retries=3, delay=1):
    """
    Decorator to retry a function upon encountering a COM error.

    Args:
        max_retries (int): Maximum number of retries.
        delay (int): Delay in seconds between retries.

    Returns:
        function: The decorated function with retry logic.
    """
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            while retries <= max_retries:
                try:
                    return func(*args, **kwargs)
                except pywintypes.com_error as e:
                    if e.args[0] == -2147418111:  # Call was rejected by callee
                        retries += 1
                        logging.warning(f"COM call to '{func.__name__}' failed with 'Call was rejected by callee', retrying ({retries}/{max_retries})...")
                        time.sleep(delay)
                    else:
                        logging.error(f"COM error in '{func.__name__}': {e}")
                        raise
                except Exception as e:
                    logging.error(f"Unexpected error in '{func.__name__}': {e}")
                    raise
            logging.critical(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
            raise Exception(f"Failed to execute '{func.__name__}' after {max_retries} retries due to COM error.")
        return wrapper
    return decorator

# ============================
# 4. Word Application Setup
# ============================

def initialize_word():
    """
    Initializes and returns the Word application object.

    Returns:
        win32com.client.CDispatch: The Word application object.
    """
    global word  # Declare 'word' as global to modify the global variable
    try:
        pythoncom.CoInitialize()
        word_app = win32com.client.Dispatch('Word.Application')
        word_app.Visible = False  # Set to True to make Word visible for debugging
        word_app.DisplayAlerts = False
        word_app.AutomationSecurity = msoAutomationSecurityForceDisable  # Disable macros
        # Removed the AutoUpdateLinks line as it's causing AttributeError
        word = word_app  # Assign to the global 'word' variable
        logging.info("Word application initialized successfully.")
        return word_app
    except Exception as e:
        logging.critical(f"Error initializing Word application: {e}")
        exit(1)  # Exit the script if Word cannot be initialized

# ============================
# 5. Constants Definition
# ============================

# Manually define necessary Word constants
msoAutomationSecurityForceDisable = 3  # Disables all macros
wdStatisticPages = 2  # ComputeStatistics for pages
wdGoToPage = 1  # What parameter for GoTo method to go to a page
wdGoToAbsolute = 1  # Which parameter for GoTo method to specify absolute positioning

# ============================
# 6. Excel Workbook Setup
# ============================

def setup_excel_workbook():
    """
    Creates and configures a new Excel workbook with the appropriate headers and column widths.

    Returns:
        Workbook: The configured Excel workbook object.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Data"
    
    # Write the header row
    header = ['Document', 'Directory', 'N# Pages', 'N# Paragraphs', 'Size (Bytes)', 'Objective', 'Scope', 'Content']
    ws.append(header)
    
    # Adjust column widths for better readability
    column_widths = [30, 100, 15, 20, 15, 50, 50, 50]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    logging.info("Excel workbook initialized with headers and column widths.")
    return wb

# ============================
# 7. Document Processing Functions
# ============================

@retry_on_COM_error(max_retries=5, delay=2)
def get_page_count(doc):
    """
    Retrieves the number of pages in the Word document.

    Args:
        doc (win32com.client.CDispatch): The Word document object.

    Returns:
        int: The number of pages.
    """
    try:
        page_count = doc.ComputeStatistics(wdStatisticPages)
        logging.debug(f"Computed page count: {page_count}")
        return page_count
    except Exception as e:
        logging.error(f"Error computing page count: {e}")
        raise

@retry_on_COM_error(max_retries=5, delay=2)
def get_paragraph_count(doc):
    """
    Retrieves the number of paragraphs in the Word document.

    Args:
        doc (win32com.client.CDispatch): The Word document object.

    Returns:
        int: The number of paragraphs.
    """
    try:
        paragraph_count = doc.Paragraphs.Count
        logging.debug(f"Computed paragraph count: {paragraph_count}")
        return paragraph_count
    except Exception as e:
        logging.error(f"Error computing paragraph count: {e}")
        raise

def get_toc_end_position(doc):
    """
    Returns the end position of the table of contents if it exists, otherwise returns 0.

    Args:
        doc (win32com.client.CDispatch): The Word document object.

    Returns:
        int: The end position of the TOC or 0 if no TOC is found.
    """
    try:
        if doc.TablesOfContents.Count > 0:
            toc = doc.TablesOfContents(1)
            toc_range = toc.Range
            logging.info(f"Table of contents found in '{doc.Name}'.")
            return toc_range.End
    except Exception as e:
        logging.error(f"Error identifying TOC in '{doc.Name}': {e}")
    return 0  # If no TOC is found, start from the beginning

def find_heading_position(doc, heading, heading_styles_set, start_pos, end_pos):
    """
    Searches for the specified heading within the document range and returns the end position of the heading if found.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading (str): The heading text to search for (case-insensitive).
        heading_styles_set (set): Set of normalized heading styles.
        start_pos (int): The starting position for the search.
        end_pos (int): The ending position for the search.

    Returns:
        int or None: The end position of the found heading. Returns None if not found.
    """
    current_pos = start_pos
    heading_lower = heading.lower()

    while current_pos < end_pos:
        para_range = doc.Range(Start=current_pos, End=end_pos)
        if para_range.Start >= para_range.End:
            break  # No more content
        para = para_range.Paragraphs(1)
        para_text = para.Range.Text.strip()
        para_style_name = para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in para_style_name.split(',')]

        # Debug
        logging.debug(f"Searching for heading '{heading}': Found paragraph '{para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Remove numbering from the heading text using the compiled pattern
            heading_text_no_number = heading_number_pattern.sub('', para_text).strip()
            heading_text_no_number_lower = heading_text_no_number.lower()

            # Compare with the target heading
            if heading_lower == heading_text_no_number_lower:
                logging.info(f"Matched heading '{heading_text_no_number}' at position {para.Range.End}.")
                return para.Range.End  # Return the end position of the heading

        # Move to the next paragraph
        current_pos = para.Range.End

    logging.warning(f"Heading '{heading}' not found between positions {start_pos} and {end_pos}.")
    return None  # Heading not found

def extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_pos):
    """
    Extracts content under a specific heading until any new heading is encountered.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        heading_end_pos (int): The position right after the heading.
        heading_styles_set (set): Set of normalized heading styles.
        end_pos (int): The end position of the search range.

    Returns:
        str: The extracted content under the heading.
    """
    content_paras = []
    current_content_pos = heading_end_pos

    while current_content_pos < end_pos:
        content_para_range = doc.Range(Start=current_content_pos, End=end_pos)
        if content_para_range.Start >= content_para_range.End:
            break  # No more content

        content_para = content_para_range.Paragraphs(1)
        content_para_text = content_para.Range.Text.strip()
        content_para_style_name = content_para.Style.NameLocal.lower()

        # Split the style name by commas to handle composite styles
        style_variants = [s.strip() for s in content_para_style_name.split(',')]

        # Debug
        logging.debug(f"Extracting content: '{content_para_text}' with style variants {style_variants}")

        # Check if any variant of the style matches the heading styles
        is_heading = any(variant in heading_styles_set for variant in style_variants)

        if is_heading:
            # Reached any new heading, stop extraction
            logging.info(f"Reached a new heading '{content_para_text}' with style variants {style_variants}. Stopping content extraction.")
            break

        # Add paragraph text to content if it's not empty
        if content_para_text:
            content_paras.append(content_para.Range.Text)

        # Move to the next paragraph
        current_content_pos = content_para.Range.End

    extracted_content = ''.join(content_paras).strip()
    logging.info(f"Extracted content under heading ending at position {heading_end_pos}: {extracted_content[:60]}...")  # Show first 60 chars for brevity
    return extracted_content

def extract_section_content(doc, headings_dict, heading_styles_set, heading_number_pattern):
    """
    Extracts and concatenates content under all specified sections based on their headings.

    Args:
        doc (win32com.client.CDispatch): The Word document object.
        headings_dict (OrderedDict): Dictionary mapping sections to their possible headings.
        heading_styles_set (set): Set of normalized heading styles.
        heading_number_pattern (re.Pattern): Compiled regex pattern for removing numbering.

    Returns:
        dict: A dictionary with section names as keys and extracted content as values.
    """
    extracted_sections = {}
    MAX_PAGE = 13
    logging.info(f"Searching limit is set to {MAX_PAGE} pages")
    try:
        logging.info(f"Entering 'extract_section_content' for sections: {list(headings_dict.keys())}")

        # Get the end position of the TOC
        toc_end = get_toc_end_position(doc)
        logging.info(f"TOC end position: {toc_end}")

        # Get the end position of page MAX_PAGE or the end of the document if less than MAX_PAGE pages
        try:
            page_max_page_range = doc.GoTo(What=wdGoToPage, Which=wdGoToAbsolute, Count=MAX_PAGE)
            end_page_max_page = page_max_page_range.End
            logging.info(f"End position of page {MAX_PAGE}: {end_page_max_page}")
        except Exception as e:
            logging.error(f"Error getting end position of max_page {MAX_PAGE}: {e}")
            end_page_max_page = doc.Content.End

        # Check if end_page_max_page is before the TOC end
        if end_page_max_page < toc_end:
            logging.info(f"End position of page {MAX_PAGE} is before the TOC end: {end_page_max_page}. Adjusting to document end.")
            end_page_max_page = doc.Content.End

        # Iterate through each section and its possible headings
        for section, headings in headings_dict.items():
            logging.info(f"Searching for section '{section}' with possible headings: {headings}")
            for heading in headings:
                heading_end_pos = find_heading_position(doc, heading, heading_styles_set, toc_end, end_page_max_page)
                if heading_end_pos:
                    # Extract content under this heading
                    extracted_content = extract_content_under_heading(doc, heading_end_pos, heading_styles_set, end_page_max_page)
                    if extracted_content:
                        # Concatenate content for the section, ensuring it's stripped
                        concatenated_content = (extracted_sections.get(section, "") + extracted_content + "\n\n").strip()
                        extracted_sections[section] = concatenated_content
                        logging.info(f"Appended content for section '{section}' with heading '{heading}'.")
                        break  # Stop searching headings for this section after the first match
            else:
                logging.warning(f"No matching heading found for section '{section}'.")

        if extracted_sections:
            logging.info(f"Extracted content for sections: {list(extracted_sections.keys())}.")
        else:
            logging.info(f"No content found for any of the specified sections.")

    except pywintypes.com_error as e:
        logging.error(f"COM error in 'extract_section_content': {e}")
    except Exception as e:
        logging.error(f"Error in 'extract_section_content': {e}")
    finally:
        logging.info(f"Exiting 'extract_section_content' for sections: {list(headings_dict.keys())}")

    return extracted_sections

def process_document(file_path, headings_dict, heading_styles_set, heading_number_pattern):
    """
    Processes a single Word document and extracts required data.

    Args:
        file_path (str): Path to the Word document.
        headings_dict (OrderedDict): Dictionary mapping sections to their possible headings.
        heading_styles_set (set): Set of normalized heading styles.
        heading_number_pattern (re.Pattern): Compiled regex pattern for removing numbering.

    Returns:
        list or None: Extracted data as a list if successful, None otherwise.
    """
    file_name = os.path.basename(file_path)
    directory_path = os.path.dirname(file_path)
    logging.info(f"Processing file: {file_path}")

    try:
        # Open the document in read-only mode without updating links
        doc = word.Documents.Open(
            FileName=file_path,
            ConfirmConversions=False,
            ReadOnly=True,
            AddToRecentFiles=False
            # Removed UpdateLinks parameter as it's not supported in Word
            # UpdateLinks=wdUpdateLinksNever
        )
        logging.info(f"Successfully opened document '{file_name}'.")
    except pywintypes.com_error as e:
        logging.error(f"COM error opening '{file_name}': {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error opening '{file_name}': {e}")
        return None

    try:
        # Get document size
        size = os.path.getsize(file_path)
        logging.info(f"Size of document: {size} bytes.")

        # Get number of pages
        pages = get_page_count(doc)
        logging.info(f"Total pages: {pages}.")

        # Get number of paragraphs
        num_paragraphs = get_paragraph_count(doc)
        logging.info(f"Number of paragraphs: {num_paragraphs}.")

        # Extract sections
        extracted_sections = extract_section_content(doc, headings_dict, heading_styles_set, heading_number_pattern)

        # Sanitize extracted content
        objective_content = remove_illegal_characters(extracted_sections.get('objective', '').strip())
        scope_content = remove_illegal_characters(extracted_sections.get('scope', '').strip())
        content_content = remove_illegal_characters(extracted_sections.get('content', '').strip())

        # Sanitize file name and directory path
        file_name_sanitized = remove_illegal_characters(file_name)
        directory_path_sanitized = remove_illegal_characters(directory_path)

        return [
            file_name_sanitized,
            directory_path_sanitized,
            pages,
            num_paragraphs,
            size,
            objective_content,
            scope_content,
            content_content
        ]

    except Exception as e:
        logging.error(f"Error processing '{file_name}': {e}")
        return None
    finally:
        # Ensure the document is closed and COM object is released
        try:
            if 'doc' in locals() and doc is not None:
                logging.info(f"Closing document '{file_name}'.")
                doc.Close(False)  # Close without saving
                del doc
                gc.collect()
                logging.info(f"Successfully closed document '{file_name}'.")
        except Exception as e:
            logging.error(f"Error closing document '{file_name}': {e}")

def write_to_excel(data, workbook):
    """
    Writes accumulated data to an Excel workbook.

    Args:
        data (list): List of data rows to write.
        workbook (openpyxl.Workbook): The Excel workbook object.
    """
    ws = workbook.active
    ws.title = "Document Data"

    # Adjust column widths for better readability
    column_widths = [30, 100, 15, 20, 15, 50, 50, 50]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Write data rows
    try:
        logging.info("Starting to write accumulated data to Excel.")
        for row in data:
            ws.append(row)
        logging.info("All accumulated data has been written to Excel.")
    except Exception as e:
        logging.error(f"Error writing data to Excel: {e}")

# ============================
# 8. Main Function
# ============================

def main():
    global word  # Declare 'word' as global to modify it within the function

    # Initialize Word application if not already initialized
    if word is None:
        initialize_word()

    # Define the heading styles to look for, ordered from highest to lowest level
    heading_styles_ordered = [
        "Heading 1", "Heading 2", "Heading 3", "Heading 4", "Heading 5", 
        "Heading 6", "Heading 7", "Heading 8", "Titre 1", "Titre 2", 
        "Titre 3", "Titre 4", "Titre 5", "Titre 6", "Titre 7", "Titre 8"
    ]

    # Convert heading styles to lowercase and create a set for faster lookup
    heading_styles_set = set(style.lower() for style in heading_styles_ordered)

    # Define the headings to search for each section
    headings_dict = OrderedDict([
        ('objective', ['Objectif', 'Objectif du document', 'But du document', 'But de ce document', 'Purpose', 'OBJECTIF ET CONTEXTE', 'CONTEXTE ET OBJECTIFS', 'Objet du document', 'OBJET', 'Introduction (commun)', 'INTRODUCTION']),
        ('scope', ['Périmètre', 'Périmètre fonctionnel']),
        ('content', ['Contenu', 'Content'])
    ])

    # Root directory to scan
    root_dir = r'C:\Users\levequer\Downloads'  # Replace with your folder path

    # List to keep track of skipped files
    skipped_files = []

    # List to accumulate data for Excel
    extracted_data_list = []

    # Counter for processed files
    file_count = 0

    # Walk through the directory and process Word documents
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if file.lower().endswith(('.doc', '.docx')):
                file_path = os.path.join(root, file)
                file_count += 1
                logging.info(f"\n--------------------------------------------------------------------------------")
                logging.info(f"Processing file {file_count}: {file_path}")
                logging.info(f"File name: {file}")
                logging.info(f"--------------------------------------------------------------------------------")

                data = process_document(file_path, headings_dict, heading_styles_set, heading_number_pattern)
                if data:
                    extracted_data_list.append(data)
                    logging.info(f"Data appended for '{file}'.")
                else:
                    skipped_files.append(file_path)

    # Initialize Excel workbook
    wb = setup_excel_workbook()

    # Write all accumulated data to Excel at once
    try:
        write_to_excel(extracted_data_list, wb)
    except Exception as e:
        logging.error(f"Error during writing to Excel: {e}")

    # Save the Excel workbook
    try:
        wb.save('doc_output.xlsx')
        logging.info("Excel workbook 'doc_output.xlsx' saved successfully.")
    except Exception as e:
        logging.error(f"Error saving Excel workbook: {e}")

    # Display the skipped files
    if skipped_files:
        logging.warning("\nThe following files were skipped due to errors:")
        for f in skipped_files:
            logging.warning(f)
    else:
        logging.info("\nAll files were processed successfully.")

# ============================
# 9. Script Entry Point
# ============================

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logging.critical(f"Unhandled exception in main: {e}")
        # The cleanup_word() function will be called automatically by atexit
    finally:
        # Ensure Word is closed
        cleanup_word()
