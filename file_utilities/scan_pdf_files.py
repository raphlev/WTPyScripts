import os
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "PDF Document Data"

# Write the header row
header = ['Document FileName', 'Document FilePath', 'Total number of pages', 'Size of document']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\SUPPORTS_FORMATIONS_INDIGO_OPALE'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith('.pdf'):
            file_path = os.path.join(root, file)

            # Increment the file counter before processing
            file_count += 1

            print(f"\nProcessing file {file_count}: {file_path}")
            print(f"File name: {file}")

            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            size = ''
            pages = ''

            try:
                # Get the size of the document
                try:
                    size = os.path.getsize(file_path)
                    print(f"Size of '{file_path}': {size} bytes")
                except Exception as e:
                    print(f"Error getting size for '{file_path}': {e}")
                    size = ''

                # Open the PDF file
                try:
                    reader = PdfReader(file_path)
                    print(f"Successfully opened PDF '{file_path}'")
                except Exception as e:
                    print(f"Error opening '{file_path}': {e}")
                    skipped_files.append(file_path)
                    continue

                # Get the total number of pages
                try:
                    pages = len(reader.pages)
                    print(f"Total pages in '{file_path}': {pages}")
                except Exception as e:
                    print(f"Error getting page count for '{file_path}': {e}")
                    pages = ''

                # Write the data to Excel
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    size
                ])
                print(f"Data written to Excel for '{file_name}'.")

            except Exception as e:
                print(f"Error processing '{file_path}': {e}")
                skipped_files.append(file_path)

print(f"\nTotal files processed: {file_count}")

# Save the Excel workbook
try:
    wb.save('pdf_output.xlsx')
    print("Excel workbook 'pdf_output.xlsx' saved successfully.")
except Exception as e:
    print(f"Error saving Excel workbook: {e}")

# Display the skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files were processed successfully.")
