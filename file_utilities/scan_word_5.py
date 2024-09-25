import os
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Ensure you have openpyxl installed:
# pip install openpyxl

# Initialize Word application
try:
    word = win32com.client.Dispatch('Word.Application')
    word.Visible = False  # Set to True temporarily for debugging
    word.DisplayAlerts = False  # Suppress Alerts in Word
except Exception as e:
    print(f"Error initializing Word application: {e}")
    exit(1)

# Constants for Word
wdStatisticPages = 2

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Document Data"

# Write the header row
header = ['Document FileName', 'Document FilePath', 'Total number of pages', 'Number of Paragraphs', 'Size of document', 'Objectif', 'Périmètre', 'Contenu']
ws.append(header)

# Adjust column widths (optional)
column_widths = [30, 100, 20, 20, 20, 50, 50, 50]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Root directory to scan
root_dir = r'C:\Users\levequer\OneDrive - TRANSITION TECHNOLOGIES PSC S.A\Documents\Backup\Projects\Safran\SAE Indigo\Technical Space\DOCUMENTATION_APPLICATIONS'  # Replace with your folder path

# List to keep track of skipped files
skipped_files = []

# Counter for processed files
file_count = 0

def extract_section_content(doc, heading_text):
    """
    Extracts content under a specific heading until the next heading of the same style.
    """
    content = ''
    try:
        # Define the range to search (start from beginning)
        rng = doc.Content
        found = rng.Find.Execute(FindText=heading_text, MatchCase=False, MatchWholeWord=True, Forward=True, Format=True)

        if found:
            start_range = rng.Duplicate
            # Move the start range to after the heading
            start_range.Collapse(0)  # 0 corresponds to wdCollapseEnd

            # Define the end range as the rest of the document
            end_range = doc.Content
            end_range.Start = start_range.Start

            # Find the next heading of the same style
            end_found = end_range.Find.Execute(FindText="*", MatchWildcards=True, MatchCase=False, MatchWholeWord=False, Forward=True, Format=True, Style=heading_style_names)

            if end_found:
                end_range.Collapse(1)  # 1 corresponds to wdCollapseStart
                start_range.End = end_range.Start
            else:
                start_range.End = doc.Content.End  # To the end of the document

            content = start_range.Text.strip()
    except Exception as e:
        print(f"Error extracting content for heading '{heading_text}': {e}")
    return content

# Define the heading styles to look for
heading_style_names = ["Heading 1", "Heading 2", "Titre 1", "Titre 2"]

for root, dirs, files in os.walk(root_dir):
    for file in files:
        if file.lower().endswith(('.doc', '.docx')):
            file_path = os.path.join(root, file)

            # Increment the file count before processing
            file_count += 1

            print(f"Processing file {file_count}: {file_path}")  # Progress message

            try:
                # Open the document in ReadOnly mode
                doc = word.Documents.Open(file_path, ReadOnly=True)
            except Exception as e:
                print(f"Error opening {file_path}: {e}")
                skipped_files.append(file_path)
                continue

            # Initialize variables with default empty values
            file_name = os.path.basename(file_path)
            size = ''
            pages = ''
            num_paragraphs = ''
            objectif_content = ''
            perimetre_content = ''
            contenu_content = ''

            try:
                # Get Size of document
                try:
                    size = os.path.getsize(file_path)
                except Exception as e:
                    print(f"Error getting size for {file_path}: {e}")
                    size = ''

                # Get Total number of pages
                try:
                    pages = doc.ComputeStatistics(wdStatisticPages)  # 2 corresponds to wdStatisticPages
                except Exception as e:
                    print(f"Error getting page count for {file_path}: {e}")
                    pages = ''

                # Get Number of Paragraphs
                try:
                    num_paragraphs = doc.Paragraphs.Count
                except Exception as e:
                    print(f"Error getting paragraph count for {file_path}: {e}")
                    num_paragraphs = ''

                # Extract content for each section
                for heading in ['Objectif', 'Périmètre', 'Contenu']:
                    content = ''
                    try:
                        # Set the Find style to look for heading styles
                        for style_name in heading_style_names:
                            word.Selection.Find.ClearFormatting()
                            word.Selection.Find.Style = doc.Styles(style_name)
                            # Use the custom function to extract content
                            content = extract_section_content(doc, heading)
                            if content:
                                break  # Found content under this style
                        if heading.lower() == 'objectif':
                            objectif_content = content
                        elif heading.lower() == 'périmètre':
                            perimetre_content = content
                        elif heading.lower() == 'contenu':
                            contenu_content = content
                    except Exception as e:
                        print(f"Error extracting content for heading '{heading}' in {file_path}: {e}")

                # Write data to Excel
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    num_paragraphs,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])

            except Exception as e:
                print(f"Error processing {file_path}: {e}")
                # Write collected data even if there was an error
                ws.append([
                    file_name,
                    file_path,
                    pages,
                    num_paragraphs,
                    size,
                    objectif_content.strip(),
                    perimetre_content.strip(),
                    contenu_content.strip()
                ])
            finally:
                # Ensure the document is closed
                try:
                    if doc is not None:
                        doc.Close(False)  # Close without saving
                except Exception as e:
                    print(f"Error closing document {file_path}: {e}")
                finally:
                    doc = None

print(f"\nTotal files processed: {file_count}")

# Close Word application
word.Quit()
word = None  # Release the COM object

# Save the Excel workbook
wb.save('output.xlsx')

# Print skipped files
if skipped_files:
    print("\nThe following files were skipped due to errors:")
    for f in skipped_files:
        print(f)
else:
    print("\nAll files processed successfully.")
