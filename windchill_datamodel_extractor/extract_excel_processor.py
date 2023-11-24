import argparse
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, NamedStyle
import csv
from extract_xml_transformer import XMLTransformer
# import pandas as pd
# from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelFileProcessor:
    def __init__(self, input_folder, output_folder, keep_csv=False):
        # Initialize the workbook creator with the directory of CSV files and the output file path
        print('-------------------------------BEGIN EXCEL PROCESSOR--------------------------------')
        self.input_folder = input_folder
        self.output_folder = output_folder
        # Define the name for the output Excel file and initialize the excel workbook
        self.output_file = os.path.join(output_folder, os.path.basename(os.path.normpath(output_folder))+'.xlsx')
        self.keep_csv = keep_csv
        self.wb = Workbook()
        self.toc = self.wb.active
        self._setup_toc()
        message = f"******  Processing input folder: {self.input_folder} ******"
        length = len(message)
        stars = '*' * length
        print(stars)
        print(stars)
        print(message)
        print(stars)
        print(stars) 

    def __del__(self):
        print('-------------------------------END  EXCEL PROCESSOR--------------------------------')
        print('-----------------------------------------------------------------------------------')
        print('-----------------------------------------------------------------------------------')
    def _setup_toc(self):
        # Set up the Table of Contents sheet
        self.toc.title = "TOC"
        # Create TOC heading style
        heading_style = NamedStyle(name="heading", font=Font(bold=True), alignment=Alignment(horizontal="center"))
        self.wb.add_named_style(heading_style)
        self.toc['A1'] = "Table of Contents"
        self.toc['A1'].style = heading_style
        self.toc.column_dimensions[get_column_letter(1)].width = 20

    def _format_worksheet(self, ws):
        # Apply bold font to the header row
        # Iterate through each row in the worksheet
        for row in ws.iter_rows():
            # Check if the row has at least two cells
            if len(row) >= 2:
                # Check if the first cell starts with "name" or "depth" and 
                # the second cell starts with "display" or "type"
                if ((row[0].value and (row[0].value.startswith("name") or row[0].value.startswith("depth"))) and
                    (row[1].value and (row[1].value.startswith("display") or row[1].value.startswith("type")))):
                    # Apply bold font to all cells in this row
                    for cell in row:
                        cell.font = Font(bold=True)

        # Adjust the column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = max_length

    def _add_csv_to_sheet(self, csv_file):
        # Add each CSV file to a new sheet in the workbook
        csv_path = os.path.join(self.output_folder, csv_file)
        
        # Solution 1 - Using CSV module from python
        # Read the CSV file and add its contents to the sheet
        try:
            with open(csv_path, mode='r', encoding='utf-8') as f:
                self._append_rows_to_sheet(csv_file,f)
        except UnicodeDecodeError:
            message = f"******  UnicodeDecodeError: file {csv_file} transcoded to latin1 instead of utf-8 ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            print(stars)            
            # If utf-8 decoding fails, try another encoding such as 'latin1'
            with open(csv_path, mode='r', encoding='latin1') as f:
                self._append_rows_to_sheet(csv_file,f)

        # Solution 2 - Using pandas - can be used if Solution 1 has perf issue .. 
        # try:
        #     # df = pd.read_csv(csv_path, sep='~', encoding='utf-8-sig')
        #     # Set keep_default_na to False to avoid treating 'None' as NaN, Specify na_values to an empty list to ensure only empty strings are treated as NaN
        #     df = pd.read_csv(csv_path, sep='~', encoding='utf-8', keep_default_na=False, na_values=[])
        #     for row in dataframe_to_rows(df, index=False, header=True):
        #         ws.append(row)
        # except UnicodeDecodeError:
        #     try:
        #         df = pd.read_csv(csv_path, sep='~', encoding='latin1', keep_default_na=False, na_values=[])
        #         for row in dataframe_to_rows(df, index=False, header=True):
        #             ws.append(row)
        #     except Exception as e:
        #         raise e
        # except pd.errors.ParserError as e:
        #     raise e   
        
        # Delete the CSV file if the keep_csv flag is False
        if not self.keep_csv:
            os.remove(csv_path)
            print(f"Deleted CSV file: {csv_path}")

    def _append_rows_to_sheet(self, csv_file, f):
        # Extract the base name without the file extension
        base_name = os.path.splitext(csv_file)[0]
        # Truncate sheet title to a maximum of 31 characters for Excel compatibility
        sheet_title = base_name[:31]
        # Check if the base name length exceeds 31 characters and log a message with hignlight
        if len(base_name) > 31:
            message = f"******  Sheet title '{base_name}' exceeds 31 characters and will be truncated to '{sheet_title}' ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            print(stars)
        # If the sheet already exists, delete it
        if sheet_title in self.wb.sheetnames:
            sheet_to_remove = self.wb[sheet_title]
            self.wb.remove(sheet_to_remove)
        # Create a new sheet
        ws = self.wb.create_sheet(title=sheet_title)
        # Read from the CSV file and append rows to the sheet
        reader = csv.reader(f, delimiter='~')
        for row in reader:
            # Check if the row contains only the string '<EMPTY_ROW>'
            if row == ['<EMPTY_ROW>']:
                # Append an empty row
                ws.append([])
            else:
                # Append the original row
                ws.append(row)
        # Apply formatting to the worksheet
        self._format_worksheet(ws)
        print(f"Added CSV file to new worksheet: {csv_file} --> {sheet_title} ")

    def _add_to_toc(self, sheet_title, index):
        # Add the sheet name to the TOC with a hyperlink
        toc_cell = self.toc.cell(column=1, row=index+1, value=sheet_title)
        # Truncate sheet title to a maximum of 31 characters for Excel compatibility
        toc_cell.hyperlink = f"#{sheet_title[:31]}!A1"
        toc_cell.style = 'Hyperlink'

    def _remove_existing_output_file(self):
        # Check if the output file already exists and remove it if it does
        try:
            if os.path.exists(self.output_file):
                os.remove(self.output_file)
                message = f"******  Deleted existing file: {self.output_file} ******"
                length = len(message)
                stars = '*' * length
                print(stars)
                print(message)
                print(stars) 
        except Exception as e:
            message = f"******  Failed to delete Excel file: {self.output_file} ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            exception_type = type(e).__name__
            print(f"{exception_type}: {e}")
            print(stars) 

    def create_excel_with_toc(self):
        # Create the workbook and save it to the output file
        self._remove_existing_output_file()
        try:
            if list(filter(lambda f: f.endswith('.csv'), os.listdir(self.output_folder))):
                index = 2  # Start index at 2
                for csv_file in sorted(os.listdir(self.output_folder)):
                    if csv_file.endswith('.csv'):
                        self._add_csv_to_sheet(csv_file)
                        self._add_to_toc(os.path.splitext(csv_file)[0], index)
                        index += 1  # Only increment index if file ends with .csv
                self.wb.save(self.output_file)
                message = f"******  Excel file saved at {self.output_file} ******"
                length = len(message)
                stars = '*' * length
                print(stars)
                print(message)
                print(stars) 
            else:
                print('No *.csv files found for : '+self.output_file+' - File not created !')
        except Exception as e:
            message = f"******  Failed to create Excel file: {self.output_file} ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            exception_type = type(e).__name__
            print(f"{exception_type}: {e}")
            print('An existing file with same name may be opened or used by someone else.')
            print(stars) 

    def process_xml_files(self):
        # Iterate each file in the input directory
        for file in os.listdir(self.input_folder):
            # Check if the file is an XML file
            if file.endswith(".xml"):
                input_file_path = os.path.join(self.input_folder, file)
                transformer = XMLTransformer(input_file_path, self.output_folder)
                # Parse and transform to CSV file
                transformer.transform()

    def create_output_directory(self):
        try:
            os.makedirs(self.output_folder, exist_ok=True)
            message = f"******  Created output directory: {self.output_folder} ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            print(stars) 
        except OSError as e:
            message = f"******  Failed to create output directory: ******"
            length = len(message)
            stars = '*' * length
            print(stars)
            print(message)
            exception_type = type(e).__name__
            print(f"{exception_type}: {e}")
            print(stars) 

    def process_excel_file(self):
        try:
            # Create the output directory if it doesn't exist
            # Check if the output directory does not exist
            if not os.path.exists(self.output_folder):
                self.create_output_directory()

            # Iterate over all files in output directory
            print("All existing *.csv files in output directory will be removed.")
            for filename in os.listdir(self.output_folder):
                if filename.endswith('.csv'):
                    os.remove(os.path.join(self.output_folder, filename))

            # Process each xml files
            self.process_xml_files()

            # Create the excel workbook
            self.create_excel_with_toc()
        except Exception as e:
            message = f"******************  Process excel file failed: ******************"
            length = len(message)
            stars = '*' * length
            marks = '!' * length
            print(stars)
            print(marks)
            print(message)
            exception_type = type(e).__name__
            print(f"{exception_type}: {e}")
            print(marks)
            print(stars)

    @staticmethod
    def run():
        # Set up command line argument parsing
        parser = argparse.ArgumentParser(description="Generate CSV files from XML Files. Generate Excel workbook from CSV files with TOC.")
        parser.add_argument('-i', '--input_dir', required=True, help='Input directory containing XML files (Types, Enum, Classification).')
        parser.add_argument('-o', '--output_dir', required=True, help='Output directory for CSV and Excel files.')
        parser.add_argument('--keep_csv', action='store_true', help='Optional: keep CSV files after processing')
        args = parser.parse_args()
        excel_processor = ExcelFileProcessor(args.input_dir, args.output_dir, args.keep_csv)
        excel_processor.process_excel_file()

if __name__ == "__main__":
    ExcelFileProcessor.run()
