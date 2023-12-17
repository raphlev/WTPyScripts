import pandas as pd
import openpyxl
import argparse
from openpyxl.styles import PatternFill
import os

def is_unique_based_on_columns(df, num_columns):
    """ Check if the DataFrame rows are unique based on the first 'num_columns' columns. """
    subset = df.iloc[:, :num_columns]
    return not subset.duplicated().any()

def compare_rows(row1, row2, num_columns):
    """ Compare rows based on the number of columns specified. """
    return all(row1.iloc[:num_columns] == row2.iloc[:num_columns])

def apply_highlighting(wb, sheet_name, row_index, col_start, fill_color):
    """ Apply highlighting to a row starting from a specific column. """
    max_col = wb[sheet_name].max_column
    for col in range(col_start, max_col + 1):
        wb[sheet_name].cell(row=row_index + 2, column=col + 1).fill = fill_color


def compare_excel_files(source_file, target_file, output_dir):
    # Load the Excel files
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    # Define colors
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    for sheet in source_wb.sheetnames:
        df_source = pd.read_excel(source_file, sheet_name=sheet, engine='openpyxl')
        df_target = pd.read_excel(target_file, sheet_name=sheet, engine='openpyxl')

        # Determine the uniqueness criterion for each sheet
        unique_by_one_column = is_unique_based_on_columns(df_source, 1)
        unique_by_two_columns = False if unique_by_one_column else is_unique_based_on_columns(df_source, 2)

        max_rows = max(len(df_source), len(df_target))
        for row in range(max_rows):
            if row < len(df_source) and row < len(df_target):
                # Compare based on the uniqueness criterion
                if (unique_by_one_column and compare_rows(df_source.iloc[row], df_target.iloc[row], 1)) or \
                   (unique_by_two_columns and compare_rows(df_source.iloc[row], df_target.iloc[row], 2)):
                    # Check for differences starting from the 4th cell
                    for col in range(3, len(df_source.columns)):
                        if row < len(df_target) and df_source.iloc[row, col] != df_target.iloc[row, col]:
                            apply_highlighting(target_wb, sheet, row, col, yellow_fill)
                else:
                    apply_highlighting(target_wb, sheet, row, 0, red_fill)
            elif row >= len(df_source):
                apply_highlighting(target_wb, sheet, row, 0, red_fill)
            elif row >= len(df_target):
                apply_highlighting(source_wb, sheet, row, 0, red_fill)

    # Save the modified workbooks
    source_output = os.path.join(output_dir, 'source_modified.xlsx')
    target_output = os.path.join(output_dir, 'target_modified.xlsx')
    source_wb.save(source_output)
    target_wb.save(target_output)

def main():
    parser = argparse.ArgumentParser(description='Compare two Excel files.')
    parser.add_argument('-s', '--source_file', required=True, help='Path to the source Excel file')
    parser.add_argument('-t', '--target_file', required=True, help='Path to the target Excel file')
    parser.add_argument('-o', '--output_dir', required=True, help='Directory to save the output files')
    
    args = parser.parse_args()
    
    compare_excel_files(args.source_file, args.target_file, args.output_dir)

if __name__ == "__main__":
    main()
