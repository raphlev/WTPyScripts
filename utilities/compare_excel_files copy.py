import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import argparse
import os

def determine_unicity(df):
    """ Determine how many columns should be considered for row unicity. """
    num_columns = df.shape[1]
    if num_columns >= 4:
        return 2  # Consider the first two columns for unicity if 4 or more columns
    else:
        return 1  # Consider only the first column for unicity if less than 4 columns

def compare_rows(row1, row2, num_columns):
    """ Compare rows based on the number of columns specified for unicity. """
    return row1.iloc[:num_columns].equals(row2.iloc[:num_columns])

def apply_highlighting(sheet, row_index, col_start, fill_color):
    """ Apply highlighting to a row starting from a specific column. """
    max_col = sheet.max_column
    for col in range(col_start, max_col + 1):
        cell = sheet.cell(row=row_index + 1, column=col)
        cell.fill = fill_color

def compare_excel_files(source_file, target_file, output_dir):
    # Load the Excel files
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    # Define colors
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    for sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[sheet_name]
        target_sheet = target_wb[sheet_name] if sheet_name in target_wb.sheetnames else None

        if target_sheet:
            # Use pandas to read the data for comparison
            df_source = pd.read_excel(source_file, sheet_name=sheet_name, engine='openpyxl')
            df_target = pd.read_excel(target_file, sheet_name=sheet_name, engine='openpyxl')

            # Determine unicity based on the number of columns
            num_columns_for_unicity = determine_unicity(df_source)

            for row_index in range(1, max(source_sheet.max_row, target_sheet.max_row) + 1):
                source_row = df_source.iloc[row_index-1] if row_index <= df_source.shape[0] else None
                target_row = df_target.iloc[row_index-1] if row_index <= df_target.shape[0] else None

                if source_row is not None and target_row is not None:
                    # Compare based on the uniqueness criterion
                    if compare_rows(source_row, target_row, num_columns_for_unicity):
                        # Check for differences starting from the 4th cell
                        for col in range(3, source_sheet.max_column):
                            if source_row[col] != target_row[col]:
                                apply_highlighting(target_sheet, row_index, col + 1, yellow_fill)
                    else:
                        apply_highlighting(target_sheet, row_index, 1, red_fill)
                elif source_row is not None:
                    apply_highlighting(source_sheet, row_index, 1, red_fill)
                elif target_row is not None:
                    apply_highlighting(target_sheet, row_index, 1, red_fill)

    # Save the modified workbooks
    source_output = os.path.join(output_dir, 'source_modified.xlsx')
    target_output = os.path.join(output_dir, 'target_modified.xlsx')
    source_wb.save(source_output)
    target_wb.save(target_output)

def main():
    parser = argparse.ArgumentParser(description='Compare two Excel files.')
    parser.add_argument('-s', '--source_file', required=True, help='Path to the source Excel file')
    parser.add_argument('-t', '--target_file', required=True, help='Path to the target Excel file')
    parser.add_argument('-o', '--output_dir', required=True, help='Directory to save the output files')
    
    args = parser.parse_args()
    
    compare_excel_files(args.source_file, args.target_file, args.output_dir)

if __name__ == "__main__":
    main()
