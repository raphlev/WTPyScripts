"""
File: hyperlink_2_url.py
Author: Tomasz Krauze
Date: January, 2024
Description: This script is designed to find corrupted sheets. Corrupted sheets need to be removed manully before changing hyperlink to url.  

"""

import openpyxl
import logging
import os
import warnings
import time
import sys

start_time = time.time()
warnings.simplefilter("ignore")
logging.basicConfig(level=logging.INFO)

def replace_hyperlinks_with_urls(worksheet, column_header):
    # Find the column index based on the header
    column_index = None
    for col in worksheet.iter_cols():
        if col[0].value == column_header:
            column_index = col[0].column
            break

    if column_index is not None:
        # Iterate through the rows in the specified column
        for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                if cell.hyperlink is not None:
                    # Get the URL from the hyperlink
                    url = cell.hyperlink.target
                    # Replace hyperlink with URL
                    cell.value = url
                    cell.hyperlink = None

# List of Excel file names to merge
folder_path = sys.argv[1]
files = os.listdir(folder_path)
excel_files = [file for file in files if file.endswith('.xlsx') or file.endswith('.xls')]

if len(sys.argv) != 2:
    print("Usage: python hyperlink_2_url.py <path>")
    print("For example: python hyperlink_2_url.py C:/Users/xxx/Desktop/excels/src/")
    print("python.exe -m pip install --upgrade pip")
    print("pip install pandas openpyxl xlsxwriter")
    sys.exit(1)

print(f"+------------------------------------------------------------+")
print(f"|                 Changing hyperlink to url                  |")
print(f"+------------------------------------------------------------+")

for file in excel_files:
    
    print(file)
    input_excel_file = os.path.join(folder_path, file)
    wb = openpyxl.load_workbook(input_excel_file, data_only=True)
    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        # Replace hyperlinks without highlighting in the specified column
        replace_hyperlinks_with_urls(ws, "P_DOC")

        # Save the modified workbook
        wb.save(input_excel_file)

end_time = time.time()
elapsed_time = round(((end_time - start_time) / 60), 2)
print(f"Total extraction time taken: {elapsed_time} minutes")